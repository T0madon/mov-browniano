import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active
# random_numbers = [random.randint(1, 100) for _ in range(20)]

# for i, num in enumerate(random_numbers, start=1):
#     sheet.cell(row=i, column=1).value = num
#     sheet.cell(row=i, column=2).value = num

plus = 1
diminuindo = 1

for i in range(1362):

    for j in range(3, 1365-(plus-1)):
        sheet.cell(row=j, column=i+7).value = f'=((C{j+plus}-C{j})^2) + ((D{j+plus}-D{j})^2)'
    plus += 1


workbook.save('./aprender.xlsx')